
import os
import csv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except Exception:
    _HAS_OPENPYXL = False

def write_test_results_excel(results, filename="reports/test_results_order.xlsx", sheet_name="Test Results"):
    if results is None:
        results = []

    # đảm bảo thư mục tồn tại (nếu có)
    dirpath = os.path.dirname(filename)
    if dirpath:
        os.makedirs(dirpath, exist_ok=True)

    headers = ["Time", "Test Name", "Keyword", "Quantity", "Actual", "Status"]

    # Nếu openpyxl có sẵn và filename là .xlsx -> ghi Excel
    if _HAS_OPENPYXL and filename.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # header
        ws.append(headers)

        # rows
        for r in results:
            ws.append([r.get(h, "") for h in headers])

        # auto width
        for i, column_cells in enumerate(ws.columns, 1):
            try:
                max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            except Exception:
                max_length = 0
            ws.column_dimensions[get_column_letter(i)].width = max_length + 2

        wb.save(filename)
        print(f"Test results saved to {os.path.abspath(filename)}")
        return os.path.abspath(filename)

    # fallback: ghi CSV (nếu filename là .xlsx thì đổi thành .csv)
    csv_path = filename
    if filename.lower().endswith(".xlsx"):
        csv_path = filename[:-5] + ".csv"

    with open(csv_path, mode="w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for r in results:
            writer.writerow([r.get(h, "") for h in headers])

    if _HAS_OPENPYXL:
        print(f"Lưu ra CSV tại {os.path.abspath(csv_path)} (openpyxl đã cài nhưng filename không .xlsx)")
    else:
        print(f"openpyxl không cài — đã ghi CSV thay thế tại {os.path.abspath(csv_path)}")
    return os.path.abspath(csv_path)
class ExcelReporter:
    def __init__(self, file_path="Reports/test_results_checkout.xlsx"):
        self.file_path = file_path
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        if not os.path.exists(file_path):
            wb = Workbook()
            ws = wb.active
            ws.title = "Checkout Results"
            ws.append([
                "Time", "Test Name", "Keyword", "Name", "Phone", "Email",
                "Province", "Address", "Expected", "Actual Error", "Status"
            ])
            wb.save(file_path)

    def write_result(self, row_data: dict):
        """Ghi kết quả test ra Excel"""
        try:
            wb = load_workbook(self.file_path)
            ws = wb.active
            ws.append([
                row_data.get("Time", ""),
                row_data.get("Test Name", ""),
                row_data.get("Keyword", ""),
                row_data.get("Name", ""),
                row_data.get("Phone", ""),
                row_data.get("Email", ""),
                row_data.get("Province", ""),
                row_data.get("Address", ""),
                row_data.get("Expected", ""),
                row_data.get("Actual Error", ""),
                row_data.get("Status", "")
            ])
            wb.save(self.file_path)
        except Exception as e:
            print(f"⚠️ Không thể ghi Excel: {e}")

REPORT_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "Reports", "addcart_report.xlsx")

def init_report():
    """Tạo file Excel nếu chưa có."""
    os.makedirs(os.path.dirname(REPORT_PATH), exist_ok=True)
    if not os.path.exists(REPORT_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = "AddCart Results"
        ws.append(["Time", "Keyword", "Quantity", "Result", "Message", "Screenshot"])
        wb.save(REPORT_PATH)

def log_result(keyword, quantity, result, message="", screenshot_path=""):
    """Ghi kết quả test vào Excel, có tô màu kết quả."""
    init_report()
    wb = load_workbook(REPORT_PATH)
    ws = wb.active

    # Màu sắc theo trạng thái
    fill_colors = {
        "PASS": "90EE90",  # xanh lá nhạt
        "FAIL": "FF7F7F",  # đỏ nhạt
        "ERROR": "FFD966"  # vàng nhạt
    }

    time_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([time_str, keyword, quantity, result, message, screenshot_path])

    # Tô màu kết quả
    last_row = ws.max_row
    color = fill_colors.get(result.upper(), "FFFFFF")
    for col in range(1, 7):
        ws.cell(row=last_row, column=col).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    wb.save(REPORT_PATH)
